#!/usr/bin/env python3
"""
exportar.py — Therian Masks CR
Convierte el Excel de inventario en productos.json para el sitio web.
Uso: python exportar.py
"""
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Instalando dependencia openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# ✅ Siempre usar la carpeta donde está este .py (no el cwd)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = os.path.join(BASE_DIR, "inventario-therian.xlsx")
JSON_FILE  = os.path.join(BASE_DIR, "productos.json")

SHEET_NAME = "Productos"
START_ROW  = 4

def exportar():
    # Debug útil (por si vuelve a fallar)
    print("📌 Carpeta del script:", BASE_DIR)
    print("📌 Directorio actual (cwd):", os.getcwd())
    print("📌 Buscando Excel en:", EXCEL_FILE)

    if not os.path.exists(EXCEL_FILE):
        print(f"❌ No encontré el archivo: {EXCEL_FILE}")
        print("✅ Asegurate de que el Excel esté en la MISMA carpeta que exportar.py")
        print("📂 Archivos en esa carpeta:")
        try:
            for f in os.listdir(BASE_DIR):
                print("   -", f)
        except Exception as e:
            print("   (No pude listar archivos)", e)
        input("Presioná Enter para salir...")
        return

    print(f"📂 Leyendo {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ No encontré la hoja '{SHEET_NAME}' en el Excel.")
        print("   Hojas disponibles:", wb.sheetnames)
        input("Presioná Enter para salir...")
        return

    ws = wb[SHEET_NAME]
    productos = []

    for row in ws.iter_rows(min_row=START_ROW, values_only=True):
        id_, nombre, desc, precio, stock_total, stock, foto, cat, dest, estado = (list(row) + [None]*10)[:10]

        if not nombre:
            continue

        if str(estado).strip().lower() != "activo":
            continue

        try:
            precio_int = int(float(str(precio).replace(",", ""))) if precio else 0
        except:
            precio_int = 0

        try:
            stock_int = int(float(str(stock))) if stock is not None else 0
        except:
            stock_int = 0

        productos.append({
            "nombre":      str(nombre).strip(),
            "descripcion": str(desc).strip() if desc else "",
            "precio":      precio_int,
            "stock":       stock_int,
            "foto":        f"imagenes/{str(foto).strip()}" if foto else "",
            "categoria":   str(cat).strip() if cat else "",
            "destacado":   str(dest).strip().upper() == "SI",
        })

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)

    print(f"✅ ¡Listo! Se exportaron {len(productos)} productos a '{JSON_FILE}'")
    print("   Abrí (o recargá) index.html en tu navegador para ver los cambios.")
    input("\nPresioná Enter para cerrar...")

if __name__ == "__main__":
    exportar()

